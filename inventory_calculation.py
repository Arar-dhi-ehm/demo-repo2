"""
Automation

inventory_calculation.py
    Capabilities:
            Read spreadsheet file and automate stuff
        Get total of inventory * price
        Write the total in the 5th column of .xlsx
        Save the xlsx programmatically
    Limitations:
    
    Improvements:

    Prerequisite:
        Install openpyxl
        
"""
import openpyxl

inventory_file = openpyxl.load_workbook('/home/renzo/Documents/git/inventory.xlsx')
# Name of tab in inventory.xlsx
product_list = inventory_file['Sheet1']

# Dictionaries:
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# range(2,_ + 1) 2 means it will start in row 2 of the file. The +1 means it will be ended in row 75 and not 74.
for product_row in range(2, product_list.max_row + 1):
    # Give supplier name for each row. (_,4) 4 means 4th column
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    # Create new column; 5th column
    inventory_price = product_list.cell(product_row, 5)

    # Calculation of product per supplier
    if supplier_name in products_per_supplier:
        # products_per_supplier['key'] = 'value'
        current_number_products = products_per_supplier.get(supplier_name)
        # Increment it by 1
        products_per_supplier[supplier_name] = current_number_products + 1
    else:
        # Will create new entry in the dict
        print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # Calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic of products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[product_number] = inventory

    # Add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# Create a new file i.e. save as
inventory_file.save('inventory_with_total_value.xlsx')