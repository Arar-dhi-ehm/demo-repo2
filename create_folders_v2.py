"""
create_folders_v2.py
    Capabilities:
        Create multiple folders from excel file data
    Limitations:
        Can't create a line for subfolders
        Can't determine which folder is a duplicate
        Will not overwrite a duplicate folder
    Improvements:

    Note:
        Spreadsheet with 3 columns of data
        
"""

import os  # For creating folder 
import openpyxl  # For reading excel files

result_location = '/home/renzo/Documents/python_prac'
spreadsheet_data = '/home/renzo/Documents/git/test_2.xlsx'

# Load the spreadsheet file
workbook = openpyxl.load_workbook(spreadsheet_data)
# Define sheet to use
sheet = workbook['Sheet1']

column_values = [(cell.value for col in sheet.iter_cols(
                min_row=2, max_row=None,min_col=1, max_col=1) for cell in col), (cell.value for col in sheet.iter_cols(
                min_row=2, max_row=None,min_col=2, max_col=2) for cell in col), (cell.value for col in sheet.iter_cols(
                min_row=2, max_row=None,min_col=3, max_col=3) for cell in col)
                ]

column_1 = column_values[0]
column_2 = column_values[1]
column_3 = column_values[2]

result = ('{}_{}_{}'.format(x, y, z) for x, y, z in zip(column_1, column_2, column_3))

# Create Folders
# for value in result:
#     print(f'Creating folder: {value}')

# Create Folders
try:
    for value in result:
        folder_name = value
        base_directory = result_location
        os.makedirs(os.path.join(base_directory, folder_name))  # Join directory and folder name
        print(f'Folder created: {folder_name}')
    print('\nDone!\n')
except FileExistsError:
    print('Error: A folder with the same name already exist.')